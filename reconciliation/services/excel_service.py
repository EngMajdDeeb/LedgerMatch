"""
Turns a raw El-Bayan (البيان) account-statement export into a clean
structured table (dict) that the AI comparison step can reason about.

This is a pure formatting/cleanup step - no semantic judgement about what a
transaction "means" happens here. It only:
  * locates the header row and maps columns by name (so column order / extra
    columns in other pharmacies' exports don't break parsing),
  * separates real transaction rows from the opening-balance row, the
    totals row and the closing-balance row,
  * returns everything as plain JSON-serializable data.

The real El-Bayan export inspected to build this (see صندوق راما.xlsx) is a
"كشف حساب" (account statement) for a pharmacy's cash account: one row per
sale invoice / return / supplier payment, with a running balance, plus a
"رصيد اول المدة" (opening balance) row at the top and a totals + closing
balance row at the bottom. Column headers observed: مدين, دائن, البيان,
الحساب, مركز الكلفة, التاريخ, رقم القيد, الرصيد.
"""
from __future__ import annotations

import datetime
import re
from dataclasses import dataclass, field

import openpyxl

# El-Bayan titles the sheet e.g. "صندوق صيدلية راما رسلان - ليرة جديدة (181112)" -
# the parenthesized number is that pharmacy's account code in El-Bayan, used
# on the combined "ميزان مراجعة" (trial balance) export (see
# excel_export_service.build_combined_results_excel).
ACCOUNT_CODE_PATTERN = re.compile(r'\((\d+)\)')

# Canonical column key -> possible header texts seen in El-Bayan exports.
COLUMN_ALIASES = {
    'debit': ['مدين'],
    'credit': ['دائن'],
    'description': ['البيان', 'بيان'],
    'account': ['الحساب', 'حساب'],
    'cost_center': ['مركز الكلفة', 'مركز التكلفة'],
    'date': ['التاريخ', 'تاريخ'],
    'entry_no': ['رقم القيد', 'رقم القيد المحاسبي'],
    'running_balance': ['الرصيد', 'رصيد'],
}

OPENING_BALANCE_MARKERS = ['رصيد اول المدة', 'رصيد أول المدة']
TOTALS_ROW_MARKER = 'مجموع'
CLOSING_BALANCE_MARKER = 'الرصيد'

MAX_HEADER_SEARCH_ROWS = 10


class ExcelParsingError(Exception):
    pass


def _cell_str(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _to_number(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    try:
        return round(float(str(value).replace(',', '')), 2)
    except (TypeError, ValueError):
        return None


def _to_iso_date(value):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if value in (None, ''):
        return None
    return str(value).strip()


def _find_header_row(ws):
    """Return (row_index, {canonical_key: col_index}) for the header row."""
    for row in ws.iter_rows(min_row=1, max_row=MAX_HEADER_SEARCH_ROWS):
        texts = {cell.column: _cell_str(cell.value) for cell in row if cell.value is not None}
        if not texts:
            continue
        col_map = {}
        for canonical, aliases in COLUMN_ALIASES.items():
            for col_idx, text in texts.items():
                if text in aliases:
                    col_map[canonical] = col_idx
                    break
        # Require at least debit/credit/description to accept this as the header row.
        if {'debit', 'credit', 'description'}.issubset(col_map.keys()):
            return row[0].row, col_map
    raise ExcelParsingError(
        'تعذر العثور على صف رؤوس الأعمدة (مدين / دائن / البيان) في ملف البيان.'
    )


def _find_period_dates(ws):
    period_from, period_to = None, None
    for row in ws.iter_rows(min_row=1, max_row=MAX_HEADER_SEARCH_ROWS):
        for cell in row:
            text = _cell_str(cell.value)
            if text in ('من تاريخ', 'من تاريخ:'):
                period_from = _to_iso_date(ws.cell(row=cell.row, column=cell.column + 1).value)
            elif text in ('الى تاريخ', 'إلى تاريخ', 'الى تاريخ:'):
                period_to = _to_iso_date(ws.cell(row=cell.row, column=cell.column + 1).value)
    return period_from, period_to


def _find_pharmacy_label(ws, transactions):
    if transactions:
        cost_center = transactions[0].get('cost_center')
        if cost_center:
            return cost_center
    # Fallback: look for the title cell (usually row 1) mentioning the box/pharmacy.
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            text = _cell_str(cell.value)
            if 'صيدلية' in text or 'صندوق' in text:
                return text
    return None


def _find_account_code(ws):
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            match = ACCOUNT_CODE_PATTERN.search(_cell_str(cell.value))
            if match:
                return match.group(1)
    return None


@dataclass
class ParsedLedger:
    pharmacy_label: str | None
    account_code: str | None
    period_from: str | None
    period_to: str | None
    opening_balance: float | None
    closing_balance: float | None
    total_debit: float | None
    total_credit: float | None
    transactions: list = field(default_factory=list)

    def to_dict(self):
        return {
            'pharmacy_label': self.pharmacy_label,
            'account_code': self.account_code,
            'period_from': self.period_from,
            'period_to': self.period_to,
            'opening_balance': self.opening_balance,
            'closing_balance': self.closing_balance,
            'total_debit': self.total_debit,
            'total_credit': self.total_credit,
            'transactions': self.transactions,
        }


def parse_el_bayan_excel(file_path_or_buffer) -> dict:
    """Parse an El-Bayan cash-account statement export into a clean dict.

    Raises ExcelParsingError if the file doesn't look like the expected
    El-Bayan ledger export.
    """
    try:
        wb = openpyxl.load_workbook(file_path_or_buffer, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface as a domain error
        raise ExcelParsingError(f'تعذر فتح ملف الإكسل: {exc}') from exc

    ws = wb.worksheets[0]
    header_row_idx, col_map = _find_header_row(ws)
    period_from, period_to = _find_period_dates(ws)

    opening_balance = None
    closing_balance = None
    total_debit_row = None
    total_credit_row = None
    transactions = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        description = _cell_str(row[col_map['description'] - 1].value) if col_map.get('description') else ''
        debit = _to_number(row[col_map['debit'] - 1].value) if col_map.get('debit') else None
        credit = _to_number(row[col_map['credit'] - 1].value) if col_map.get('credit') else None
        entry_no = _cell_str(row[col_map['entry_no'] - 1].value) if col_map.get('entry_no') else ''
        running_balance = _to_number(row[col_map['running_balance'] - 1].value) if col_map.get('running_balance') else None

        if not description and debit is None and credit is None:
            continue  # blank row

        if any(marker in description for marker in OPENING_BALANCE_MARKERS):
            opening_balance = running_balance if running_balance is not None else (debit or credit)
            continue

        if TOTALS_ROW_MARKER in description:
            total_debit_row = debit
            total_credit_row = credit
            continue

        if entry_no:
            transactions.append({
                'debit': debit,
                'credit': credit,
                'description': description,
                'account': _cell_str(row[col_map['account'] - 1].value) if col_map.get('account') else '',
                'cost_center': _cell_str(row[col_map['cost_center'] - 1].value) if col_map.get('cost_center') else '',
                'date': _to_iso_date(row[col_map['date'] - 1].value) if col_map.get('date') else None,
                'entry_no': entry_no,
                'running_balance': running_balance,
            })
            continue

        if CLOSING_BALANCE_MARKER in description:
            closing_balance = running_balance if running_balance is not None else (debit or credit)
            continue

    total_debit = total_debit_row if total_debit_row is not None else round(
        sum(t['debit'] for t in transactions if t['debit']), 2
    )
    total_credit = total_credit_row if total_credit_row is not None else round(
        sum(t['credit'] for t in transactions if t['credit']), 2
    )

    if closing_balance is None:
        # Fall back to the running balance of the last transaction row.
        for t in reversed(transactions):
            if t['running_balance'] is not None:
                closing_balance = t['running_balance']
                break

    parsed = ParsedLedger(
        pharmacy_label=_find_pharmacy_label(ws, transactions),
        account_code=_find_account_code(ws),
        period_from=period_from,
        period_to=period_to,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        total_debit=total_debit,
        total_credit=total_credit,
        transactions=transactions,
    )

    if not transactions:
        raise ExcelParsingError('لم يتم العثور على أي حركات (قيود) في ملف البيان.')

    return parsed.to_dict()
