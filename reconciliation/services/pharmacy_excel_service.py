"""
Alternative to vision_service.py: when a pharmacy submits its daily
cash-box sheet as a real spreadsheet instead of a photo, this reads it
deterministically - no OCR guesswork, so it's strictly more reliable than
the vision path whenever it's available.

Returns the exact same shape vision_service.extract_image_table does
(pharmacy_name / report_date / day_totals / shifts / notes / row_notes), so
aggregation.py and comparison_service.py don't need to know or care which
path produced it.

The label column (one row per category: مبيعات، مرتجع مبيعات...) is found
by *content* - which column actually contains recognizable category
names - rather than by requiring a specific header word like "الاسم". A
first version required that exact header and broke on real pharmacy
files that don't use it; matching by content is robust to whatever header
text (or none at all) a given sheet uses for that column.

Shift columns (one per work shift, e.g. الكادر الصباحي) and the notes
column are still found by header text ("الكادر"/"وردية", "ملاحظ"), with a
fallback to treating any other populated header column as an unlabeled
data column so a sheet with a single unnamed values column still parses.
"""
from __future__ import annotations

import datetime

import openpyxl

from ..constants import ALL_CATEGORY_KEYS, BALANCE_CATEGORY_KEYS, CATEGORY_LABELS_AR, FLOW_CATEGORY_KEYS

MAX_META_SEARCH_ROWS = 15
MIN_LABEL_MATCHES = 3
HEADER_LOOKBACK_ROWS = 5

_LABEL_LOOKUP = {''.join(label.split()): key for key, label in CATEGORY_LABELS_AR.items()}


class PharmacyExcelParsingError(Exception):
    pass


def _cell_str(value) -> str:
    return '' if value is None else str(value).strip()


def _to_number(value):
    if value is None or value == '' or value == '-':
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    try:
        return round(float(str(value).replace(',', '')), 2)
    except (TypeError, ValueError):
        return None


def _match_category(text: str):
    normalized = ''.join(text.split())
    if not normalized:
        return None
    if normalized in _LABEL_LOOKUP:
        return _LABEL_LOOKUP[normalized]
    for label_normalized, key in _LABEL_LOOKUP.items():
        if label_normalized and label_normalized in normalized:
            return key
    return None


def _find_label_column(ws):
    """Find the column holding category row labels by content: whichever
    column has the most cells matching a known category name. Real sheets
    vary in what (if anything) they title this column, so header text isn't
    a reliable signal here - the row contents are.
    """
    best_col = None
    best_matches = {}
    for col in range(1, ws.max_column + 1):
        matches = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell = row[col - 1]
            text = _cell_str(cell.value)
            if not text:
                continue
            key = _match_category(text)
            if key and key not in matches:
                matches[key] = cell.row
        if len(matches) > len(best_matches):
            best_matches = matches
            best_col = col

    if best_col is None or len(best_matches) < MIN_LABEL_MATCHES:
        raise PharmacyExcelParsingError(
            f'تعذر التعرف على عمود بنود الصندوق (مثل مبيعات، مرتجع مبيعات، صندوق نهاية...) '
            f'في ورقة "{ws.title}" من ملف الإكسل. تأكد أن هذه هي الورقة الصحيحة لهذه الصيدلية '
            f'وأن أسماء البنود مكتوبة بنفس صيغتها المعتادة.'
        )
    return best_col, min(best_matches.values())


def _find_header_row(ws, label_col, first_data_row):
    """Row holding the shift/notes column headers, used only to classify
    the *other* columns - not the label column, which is already known.
    """
    search_start = max(1, first_data_row - HEADER_LOOKBACK_ROWS)
    for row in ws.iter_rows(min_row=search_start, max_row=max(search_start, first_data_row - 1)):
        for cell in row:
            if cell.column == label_col:
                continue
            text = _cell_str(cell.value)
            if 'الكادر' in text or 'وردية' in text:
                return row[0].row
    # No explicit shift header found - the row directly above the first
    # recognized category row is the best remaining guess.
    return max(1, first_data_row - 1)


def _find_columns(ws, header_row_idx, label_col):
    shift_cols = []
    notes_col = None
    generic_cols = []
    for col in range(1, ws.max_column + 1):
        if col == label_col:
            continue
        text = _cell_str(ws.cell(row=header_row_idx, column=col).value)
        if not text:
            continue
        if 'ملاحظ' in text:
            notes_col = col
        elif 'الكادر' in text or 'وردية' in text:
            shift_cols.append((col, text))
        else:
            generic_cols.append((col, text))

    if not shift_cols:
        # No explicit shift headers - fall back to any other populated,
        # non-notes header column as a single unlabeled data column, rather
        # than failing outright.
        shift_cols = [c for c in generic_cols if c[0] != notes_col]

    if notes_col is None:
        claimed = {col for col, _ in shift_cols}
        for col in range(ws.max_column, label_col, -1):
            if col in claimed:
                continue
            if _cell_str(ws.cell(row=header_row_idx, column=col).value):
                notes_col = col
                break

    return shift_cols, notes_col


def _sheet_name_words(ws):
    """Words to match a sheet against a pharmacy name: the sheet's own tab
    title plus whatever text sits in its first couple of rows (where a
    per-pharmacy sheet usually repeats the full pharmacy name, e.g. sheet
    tab "راما" with A1 "صيدلية راما رسلان").
    """
    texts = [ws.title]
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            text = _cell_str(cell.value)
            if text:
                texts.append(text)
    words = set()
    for text in texts:
        words.update(text.replace('صيدلية', ' ').split())
    return words


def _select_sheet(wb, pharmacy_name: str | None):
    """Pick the sheet that belongs to this pharmacy.

    Real El-Bayan/pharmacy workbooks are often one file covering every
    pharmacy for the day: a summary roll-up sheet, sometimes a trial-balance
    sheet, and then one sheet per pharmacy holding the actual cash-box
    layout. Blindly reading the first sheet picks up the roll-up instead -
    this matches the sheet whose title/heading text overlaps the pharmacy's
    name the most.
    """
    if len(wb.worksheets) == 1 or not pharmacy_name:
        return wb.worksheets[0]

    pharmacy_words = {w for w in pharmacy_name.replace('صيدلية', ' ').split() if len(w) > 1}
    if not pharmacy_words:
        return wb.worksheets[0]

    best_sheet, best_score = None, 0
    for ws in wb.worksheets:
        sheet_words = _sheet_name_words(ws)
        score = len(pharmacy_words & sheet_words)
        if any(w in ws.title for w in pharmacy_words):
            score += 1
        if score > best_score:
            best_sheet, best_score = ws, score

    if best_sheet is None:
        available = '، '.join(ws.title for ws in wb.worksheets)
        raise PharmacyExcelParsingError(
            f'يحتوي ملف الإكسل على عدة أوراق ولم يتم العثور على ورقة تخص "{pharmacy_name}". '
            f'الأوراق الموجودة: {available}.'
        )
    return best_sheet


def _find_pharmacy_meta(ws):
    pharmacy_name = None
    report_date = None
    for row in ws.iter_rows(min_row=1, max_row=MAX_META_SEARCH_ROWS):
        for cell in row:
            value = cell.value
            text = _cell_str(value)
            if pharmacy_name is None and 'صيدلية' in text:
                pharmacy_name = text
            if report_date is None and isinstance(value, (datetime.datetime, datetime.date)):
                report_date = (
                    value.date().isoformat() if isinstance(value, datetime.datetime) else value.isoformat()
                )
    return pharmacy_name, report_date


def parse_pharmacy_excel(file_path_or_buffer, pharmacy_name: str | None = None) -> dict:
    """Parse a pharmacy-submitted cash-box .xlsx.

    `pharmacy_name` (the Django Pharmacy record's name) is used to pick the
    right sheet when the workbook covers multiple pharmacies - see
    _select_sheet. Safe to omit for a single-sheet file.

    Raises PharmacyExcelParsingError if the sheet doesn't look like the
    expected cash-box layout (no recognizable category rows, or no data
    columns at all).
    """
    try:
        wb = openpyxl.load_workbook(file_path_or_buffer, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface as a domain error
        raise PharmacyExcelParsingError(f'تعذر فتح ملف الإكسل: {exc}') from exc

    ws = _select_sheet(wb, pharmacy_name)
    label_col, first_data_row = _find_label_column(ws)
    header_row_idx = _find_header_row(ws, label_col, first_data_row)
    shift_cols, notes_col = _find_columns(ws, header_row_idx, label_col)
    pharmacy_name, report_date = _find_pharmacy_meta(ws)

    if not shift_cols:
        raise PharmacyExcelParsingError(
            'تعذر العثور على أي عمود بيانات (مثل "الكادر الصباحي") بجانب عمود أسماء البنود في ملف الصندوق.'
        )

    shifts = [
        {'shift_label': text, 'hours': None, 'values': {key: None for key in FLOW_CATEGORY_KEYS}}
        for _, text in shift_cols
    ]
    day_totals = {key: None for key in BALANCE_CATEGORY_KEYS}
    row_notes = {key: None for key in ALL_CATEGORY_KEYS}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label_text = _cell_str(row[label_col - 1].value)
        if not label_text:
            continue
        key = _match_category(label_text)
        if key is None:
            continue
        notes_col_consumed_as_value = False

        if key in FLOW_CATEGORY_KEYS:
            for idx, (col, _label) in enumerate(shift_cols):
                shifts[idx]['values'][key] = _to_number(row[col - 1].value)
        else:
            # Day-level balance: a single figure for the whole day, wherever
            # it was written - check the shift columns first, then notes.
            value = None
            for col, _label in shift_cols:
                value = _to_number(row[col - 1].value)
                if value is not None:
                    break
            if value is None and notes_col:
                value = _to_number(row[notes_col - 1].value)
                notes_col_consumed_as_value = value is not None
            day_totals[key] = value

        # Only treat the notes-column cell as an actual annotation if it
        # wasn't just consumed above as a balance figure that happened to be
        # written under that column (a known layout quirk on some sheets).
        if notes_col and not notes_col_consumed_as_value:
            note_text = _cell_str(row[notes_col - 1].value)
            if note_text and note_text != '-':
                row_notes[key] = note_text

    return {
        'pharmacy_name': pharmacy_name,
        'report_date': report_date,
        'day_totals': day_totals,
        'shifts': shifts,
        'notes': None,
        'row_notes': row_notes,
    }
