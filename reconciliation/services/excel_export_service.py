"""
Builds a real, downloadable .xlsx that mirrors the result card shown in the
web UI - same status banner, same stat boxes, same category table with the
same colour coding - so opening the file feels like the page, not like a
raw data dump. A second sheet keeps the original per-shift breakdown of
whatever was extracted from the pharmacy's submission (photo or excel), for
deeper spot-checking.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..constants import BALANCE_CATEGORY_KEYS, CATEGORY_LABELS_AR, FLOW_CATEGORY_KEYS

# Colours lifted straight from static/css/app.css so the file reads as the
# same product as the dashboard, not a generic spreadsheet export.
TEXT = '1F2733'
TEXT_MUTED = '6B7480'
BORDER_COLOR = 'DDE2EA'
GREEN = '0F9D58'
GREEN_BG = 'E6F6ED'
RED = 'D93025'
RED_BG = 'FDECEA'
AMBER = 'B06A00'
AMBER_BG = 'FEF3E2'
HEADER_BG = 'F3F5F8'

TITLE_FONT = Font(name='Calibri', size=16, bold=True, color=TEXT)
BANNER_FONT = Font(name='Calibri', size=12, bold=True)
LABEL_FONT = Font(name='Calibri', size=9, bold=True, color=TEXT_MUTED)
VALUE_FONT = Font(name='Calibri', size=13, bold=True, color=TEXT)
BODY_FONT = Font(name='Calibri', size=11, color=TEXT)
BODY_BOLD_FONT = Font(name='Calibri', size=11, bold=True, color=TEXT)
MUTED_FONT = Font(name='Calibri', size=10, italic=True, color=TEXT_MUTED)
TABLE_HEADER_FONT = Font(name='Calibri', size=10, bold=True, color=TEXT_MUTED)

THIN_BORDER = Border(bottom=Side(style='thin', color=BORDER_COLOR))
BOX_BORDER = Border(*[Side(style='thin', color=BORDER_COLOR)] * 4)

STATUS_STYLES = {
    'match': (GREEN, GREEN_BG, 'مطابق'),
    'mismatch': (RED, RED_BG, 'فرق قيمة'),
    'missing_in_bayan': (AMBER, AMBER_BG, 'غير موجود في البيان'),
    'missing_in_image': (AMBER, AMBER_BG, 'غير موجود في الملف'),
    'not_applicable': (TEXT_MUTED, 'FFFFFF', 'لا يوجد بيانات'),
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=f'FF{hex_color}', end_color=f'FF{hex_color}', fill_type='solid')


def _fmt(value):
    if value is None:
        return '-'
    return value


def _build_comparison_sheet(ws, pharmacy_name, comparison: dict):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    for col, width in zip('ABCDEF', (26, 16, 16, 14, 20, 30)):
        ws.column_dimensions[col].width = width

    row = 1
    ws.cell(row=row, column=1, value=pharmacy_name or 'الصيدلية').font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    has_differences = comparison.get('has_differences', False)
    status_color, status_bg, _ = ('D93025', 'FDECEA', None) if has_differences else ('0F9D58', 'E6F6ED', None)
    status_text = 'يوجد فروقات' if has_differences else 'مطابق تماما'
    banner = ws.cell(row=row, column=1, value=f'  {status_text}  ')
    banner.font = Font(name='Calibri', size=11, bold=True, color=status_color)
    banner.fill = _fill(status_bg)
    banner.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 2

    summary = comparison.get('summary_ar')
    if summary:
        cell = ws.cell(row=row, column=1, value=summary)
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 34
        row += 2

    totals = comparison.get('totals_diff') or {}
    stat_boxes = [
        ('صندوق نهاية / البيان', totals.get('bayan_closing_balance'), TEXT),
        ('صندوق نهاية / الملف', totals.get('image_closing_balance'), TEXT),
        ('الفرق', totals.get('difference'), RED if has_differences else GREEN),
    ]
    label_row, value_row = row, row + 1
    for idx, (label, value, color) in enumerate(stat_boxes):
        col = idx + 1
        lc = ws.cell(row=label_row, column=col, value=label)
        lc.font = LABEL_FONT
        lc.alignment = Alignment(horizontal='center')
        lc.fill = _fill(HEADER_BG)
        lc.border = BOX_BORDER
        vc = ws.cell(row=value_row, column=col, value=_fmt(value))
        vc.font = Font(name='Calibri', size=13, bold=True, color=color)
        vc.alignment = Alignment(horizontal='center')
        vc.fill = _fill(HEADER_BG)
        vc.border = BOX_BORDER
    ws.row_dimensions[label_row].height = 16
    ws.row_dimensions[value_row].height = 20
    row = value_row + 2

    headers = ['البند', 'قيمة البيان', 'قيمة الملف', 'الفرق', 'الحالة', 'ملاحظات الصيدلية']
    header_row = row
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = TABLE_HEADER_FONT
        cell.fill = _fill(HEADER_BG)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
    row += 1

    for category in comparison.get('categories', []):
        status = category.get('status', 'not_applicable')
        color, bg, status_label = STATUS_STYLES.get(status, STATUS_STYLES['not_applicable'])
        values = [
            category.get('label_ar', ''),
            _fmt(category.get('bayan_value')),
            _fmt(category.get('image_value')),
            _fmt(category.get('difference')),
            status_label,
            category.get('image_note') or '-',
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right', wrap_text=(col == 6))
            if col == 4 and status == 'mismatch':
                cell.font = Font(name='Calibri', size=11, bold=True, color=RED)
            elif col == 5:
                cell.font = Font(name='Calibri', size=10, bold=True, color=color)
            elif col == 6:
                cell.font = MUTED_FONT
            else:
                cell.font = BODY_FONT
            if status != 'not_applicable':
                cell.fill = _fill(bg)
        row += 1

    row += 1
    footer = ws.cell(
        row=row, column=1,
        value=f"{comparison.get('matched_rows_count', 0)} بند مطابق · {comparison.get('mismatched_rows_count', 0)} بند مختلف",
    )
    footer.font = MUTED_FONT


def _build_source_detail_sheet(ws, image_extraction: dict):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    shifts = image_extraction.get('shifts') or []
    day_totals = image_extraction.get('day_totals') or {}

    ws.cell(row=1, column=1, value='الصيدلية').font = BODY_BOLD_FONT
    ws.cell(row=1, column=2, value=image_extraction.get('pharmacy_name') or '')
    ws.cell(row=2, column=1, value='تاريخ التقرير').font = BODY_BOLD_FONT
    ws.cell(row=2, column=2, value=image_extraction.get('report_date') or '')

    balance_header_row = 4
    ws.cell(row=balance_header_row, column=1, value='بنود اليوم كاملا').font = TABLE_HEADER_FONT
    ws.cell(row=balance_header_row, column=1).fill = _fill(HEADER_BG)
    ws.cell(row=balance_header_row, column=2, value='القيمة').font = TABLE_HEADER_FONT
    ws.cell(row=balance_header_row, column=2).fill = _fill(HEADER_BG)

    row_idx = balance_header_row + 1
    for key in BALANCE_CATEGORY_KEYS:
        ws.cell(row=row_idx, column=1, value=CATEGORY_LABELS_AR[key]).font = BODY_BOLD_FONT
        ws.cell(row=row_idx, column=2, value=_fmt(day_totals.get(key)))
        row_idx += 1

    flow_header_row = row_idx + 1
    ws.cell(row=flow_header_row, column=1, value='بنود حسب الوردية').font = TABLE_HEADER_FONT
    ws.cell(row=flow_header_row, column=1).fill = _fill(HEADER_BG)

    for idx, shift in enumerate(shifts):
        col = idx + 2
        label = shift.get('shift_label') or f'وردية {idx + 1}'
        hours = shift.get('hours') or ''
        cell = ws.cell(row=flow_header_row, column=col, value=f'{label}\n{hours}'.strip())
        cell.font = TABLE_HEADER_FONT
        cell.fill = _fill(HEADER_BG)
        cell.alignment = Alignment(wrap_text=True, horizontal='center')

    total_col = len(shifts) + 2
    total_cell = ws.cell(row=flow_header_row, column=total_col, value='الإجمالي')
    total_cell.font = TABLE_HEADER_FONT
    total_cell.fill = _fill(HEADER_BG)

    for offset, key in enumerate(FLOW_CATEGORY_KEYS):
        r = flow_header_row + 1 + offset
        ws.cell(row=r, column=1, value=CATEGORY_LABELS_AR[key]).font = BODY_BOLD_FONT
        row_total = 0
        has_value = False
        for idx, shift in enumerate(shifts):
            col = idx + 2
            value = shift.get('values', {}).get(key)
            ws.cell(row=r, column=col, value=_fmt(value))
            if value is not None:
                row_total += value
                has_value = True
        ws.cell(row=r, column=total_col, value=round(row_total, 2) if has_value else '-').font = BODY_BOLD_FONT

    notes = image_extraction.get('notes')
    if notes:
        notes_row = flow_header_row + len(FLOW_CATEGORY_KEYS) + 2
        ws.cell(row=notes_row, column=1, value='ملاحظات').font = BODY_BOLD_FONT
        ws.cell(row=notes_row, column=2, value=notes)

    ws.column_dimensions[get_column_letter(1)].width = 22
    for col in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _round_or_none(value, ndigits=0):
    return None if value is None else round(value, ndigits)


def build_combined_results_excel(selected_date, rows: list) -> io.BytesIO:
    """Build the "ميزان مراجعة" (trial balance) workbook combining every
    pharmacy's closing-balance comparison for one day into a single sheet -
    the same layout this business already keeps by hand (رقم / الاسم /
    رصيد بيان / رصيد تقرير / الفرق / الفرق ع ق / الملاحظات), generated
    from the app's own comparison results instead of being typed up
    afterwards.

    `rows` is a list of dicts: {account_code, pharmacy_name, bayan_balance,
    report_balance, note}. Difference and its piastre (قرش) equivalent are
    computed here, not passed in, so they can never drift from the two
    balances shown next to them.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'ميزان مراجعة'
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    widths = {'A': 8, 'B': 34, 'C': 14, 'D': 14, 'E': 14, 'F': 16, 'G': 22, 'H': 3, 'I': 34, 'J': 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.cell(row=1, column=1, value='ميزان مراجعة').font = TITLE_FONT
    ws.cell(row=1, column=2, value='صناديق صيدليات ليرة جديدة').font = BODY_BOLD_FONT
    ws.cell(row=1, column=3, value='من تاريخ').font = LABEL_FONT
    date_cell = ws.cell(row=1, column=4, value=selected_date)
    date_cell.font = BODY_BOLD_FONT
    date_cell.number_format = 'mm-dd-yy'

    header_row = 3
    headers = ['الرقم', 'الاسم', 'رصيد بيان', 'رصيد تقرير', 'الفرق', 'الفرق ع ق', 'الملاحظات']
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = TABLE_HEADER_FONT
        cell.fill = _fill(HEADER_BG)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    side_list_total = 0.0
    side_row = header_row + 1

    row = header_row + 1
    for entry in rows:
        bayan = entry.get('bayan_balance')
        report = entry.get('report_balance')
        difference = None if bayan is None or report is None else round(bayan - report, 2)
        difference_piastres = None if difference is None else round(difference * 100)
        has_diff = bool(difference and abs(difference) >= 0.01)

        values = [
            entry.get('account_code') or '',
            entry.get('pharmacy_name') or '',
            _round_or_none(bayan),
            _round_or_none(report),
            _round_or_none(difference),
            difference_piastres if difference_piastres is not None else '-',
            entry.get('note') or '',
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=_fmt(value))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if col in (1, 3, 4, 5, 6) else 'right')
            if col in (5, 6) and has_diff:
                cell.font = Font(name='Calibri', size=11, bold=True, color=RED)
            elif col == 2:
                cell.font = BODY_BOLD_FONT
            else:
                cell.font = BODY_FONT
        row += 1

        if bayan is not None:
            side_list_total += bayan
            ws.cell(row=side_row, column=9, value=f"صندوق {entry.get('pharmacy_name') or ''} - ليرة جديدة").font = BODY_FONT
            ws.cell(row=side_row, column=10, value=round(bayan, 2)).font = BODY_FONT
            side_row += 1

    if side_row > header_row + 1:
        ws.cell(row=header_row, column=9, value='صناديق صيدليات ليرة جديدة').font = TABLE_HEADER_FONT
        total_cell = ws.cell(row=header_row, column=10, value=round(side_list_total, 2))
        total_cell.font = BODY_BOLD_FONT

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_audit_excel(image_extraction: dict, comparison: dict | None = None) -> io.BytesIO:
    wb = Workbook()

    comparison_sheet = wb.active
    if comparison:
        comparison_sheet.title = 'المقارنة'
        _build_comparison_sheet(comparison_sheet, image_extraction.get('pharmacy_name'), comparison)
        detail_sheet = wb.create_sheet('تفاصيل ملف الصيدلية')
    else:
        detail_sheet = comparison_sheet

    detail_sheet.title = 'تفاصيل ملف الصيدلية' if comparison else 'صندوق - مستخرج من الملف'
    _build_source_detail_sheet(detail_sheet, image_extraction)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
